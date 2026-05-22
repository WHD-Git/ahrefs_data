"""
Ahrefs keyword rankings change report — URL / page-level analysis

Input:
- CSV or xlsx named "ahrefs_site_rank_change" in the directory set by env var "ahrefs_dir"

Expected columns (case-insensitive):
  Keyword, Volume, Previous organic traffic, Current organic traffic,
  Previous position, Current position, Previous URL, Current URL

Output:
- "url_report.xlsx" in the same directory, with two sheets:
  1) URL Report  — one row per page, previous / current / delta metrics
  2) Highlights  — key winners and losers
"""

import os
from urllib.parse import urlparse

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ENV_DIR_KEY = "ahrefs_dir"
INPUT_BASENAME = "ahrefs_site_rank_change"
OUTPUT_FILENAME = "ahrefs_site_rank_change_report.xlsx"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def resolve_input_path(base_dir: str, basename: str) -> str:
    for ext in ("", ".csv", ".xlsx", ".xls"):
        p = os.path.join(base_dir, basename + ext)
        if os.path.isfile(p):
            return p
    return os.path.join(base_dir, basename + ".csv")


def detect_encoding(path: str) -> str:
    """Detect encoding from BOM bytes; falls back to utf-8."""
    with open(path, "rb") as f:
        bom = f.read(4)
    if bom[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    if bom[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    return "utf-8"


def find_col(col_map: dict, name: str) -> str:
    col = col_map.get(name.strip().lower())
    if col is None:
        raise KeyError(f'Required column "{name}" not found in the input file.')
    return col


def normalize_url(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none", "") else s.lower().rstrip("/")


def coerce_position(val):
    if val is None:
        return None
    try:
        f = float(str(val).strip())
        if pd.isna(f):
            return None
        i = int(f)
        return i if i > 0 else None
    except Exception:
        return None


def is_homepage(url: str) -> bool:
    s = url if "://" in url else "https://" + url
    try:
        return urlparse(s).path.rstrip("/") == ""
    except Exception:
        return False


def pos_le(pos_series: pd.Series, n: int) -> pd.Series:
    return pos_series.apply(lambda x: x is not None and 1 <= x <= n)


# ---------------------------------------------------------------------------
# URL Report builder
# ---------------------------------------------------------------------------

def build_url_report(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {str(c).strip().lower(): c for c in df.columns}

    kw_col        = find_col(col_map, "Keyword")
    vol_col       = find_col(col_map, "Volume")
    prev_tr_col   = find_col(col_map, "Previous organic traffic")
    curr_tr_col   = find_col(col_map, "Current organic traffic")
    prev_pos_col  = find_col(col_map, "Previous position")
    curr_pos_col  = find_col(col_map, "Current position")
    prev_url_col  = find_col(col_map, "Previous URL")
    curr_url_col  = find_col(col_map, "Current URL")

    prev_url = df[prev_url_col].apply(normalize_url)
    curr_url = df[curr_url_col].apply(normalize_url)
    prev_pos = df[prev_pos_col].apply(coerce_position)
    curr_pos = df[curr_pos_col].apply(coerce_position)
    vol      = pd.to_numeric(df[vol_col], errors="coerce").fillna(0)
    prev_tr  = pd.to_numeric(df[prev_tr_col], errors="coerce").fillna(0)
    curr_tr  = pd.to_numeric(df[curr_tr_col], errors="coerce").fillna(0)

    prev_t3  = pos_le(prev_pos, 3)
    prev_t10 = pos_le(prev_pos, 10)
    prev_t20 = pos_le(prev_pos, 20)
    curr_t3  = pos_le(curr_pos, 3)
    curr_t10 = pos_le(curr_pos, 10)
    curr_t20 = pos_le(curr_pos, 20)

    all_urls = (set(prev_url[prev_url != ""]) | set(curr_url[curr_url != ""]))

    rows = []
    for url in sorted(all_urls):
        pm = prev_url == url
        cm = curr_url == url

        p_kw  = int(pm.sum())
        c_kw  = int(cm.sum())
        p_t3  = int((pm & prev_t3).sum())
        c_t3  = int((cm & curr_t3).sum())
        p_t10 = int((pm & prev_t10).sum())
        c_t10 = int((cm & curr_t10).sum())
        p_t20 = int((pm & prev_t20).sum())
        c_t20 = int((cm & curr_t20).sum())

        p_tr = int(prev_tr[pm].sum())
        c_tr = int(curr_tr[cm].sum())

        p_vol    = int(vol[pm].sum())
        c_vol    = int(vol[cm].sum())
        p_vol_t3  = int(vol[pm & prev_t3].sum())
        c_vol_t3  = int(vol[cm & curr_t3].sum())
        p_vol_t10 = int(vol[pm & prev_t10].sum())
        c_vol_t10 = int(vol[cm & curr_t10].sum())
        p_vol_t20 = int(vol[pm & prev_t20].sum())
        c_vol_t20 = int(vol[cm & curr_t20].sum())

        new_rankings  = int((cm & ~pm).sum())
        lost_rankings = int((pm & ~cm).sum())

        # Ranking Score: Top 3 = 3pts, Top 4-10 = 2pts, Top 11-20 = 1pt
        p_score = p_t3 * 3 + (p_t10 - p_t3) * 2 + (p_t20 - p_t10) * 1
        c_score = c_t3 * 3 + (c_t10 - c_t3) * 2 + (c_t20 - c_t10) * 1

        rows.append({
            "URL":              url,
            "Is Homepage":      is_homepage(url),
            # Ranking Score
            "Prev Score":       p_score,
            "Curr Score":       c_score,
            "Score Delta":      c_score - p_score,
            # Keywords
            "Prev KWs":         p_kw,
            "Curr KWs":         c_kw,
            "KW Delta":         c_kw - p_kw,
            # Top 3
            "Prev Top 3":       p_t3,
            "Curr Top 3":       c_t3,
            "Top 3 Delta":      c_t3 - p_t3,
            # Top 10
            "Prev Top 10":      p_t10,
            "Curr Top 10":      c_t10,
            "Top 10 Delta":     c_t10 - p_t10,
            # Top 20
            "Prev Top 20":      p_t20,
            "Curr Top 20":      c_t20,
            "Top 20 Delta":     c_t20 - p_t20,
            # Traffic
            "Prev Traffic":     p_tr,
            "Curr Traffic":     c_tr,
            "Traffic Delta":    c_tr - p_tr,
            # Volume (all KWs)
            "Prev Volume":      p_vol,
            "Curr Volume":      c_vol,
            "Volume Delta":     c_vol - p_vol,
            # Volume (Top 3 KWs)
            "Prev Top 3 Vol":   p_vol_t3,
            "Curr Top 3 Vol":   c_vol_t3,
            "Top 3 Vol Delta":  c_vol_t3 - p_vol_t3,
            # Volume (Top 10 KWs)
            "Prev Top 10 Vol":  p_vol_t10,
            "Curr Top 10 Vol":  c_vol_t10,
            "Top 10 Vol Delta": c_vol_t10 - p_vol_t10,
            # Volume (Top 20 KWs)
            "Prev Top 20 Vol":  p_vol_t20,
            "Curr Top 20 Vol":  c_vol_t20,
            "Top 20 Vol Delta": c_vol_t20 - p_vol_t20,
            # Movement
            "New Rankings":     new_rankings,
            "Lost Rankings":    lost_rankings,
        })

    result = pd.DataFrame(rows)
    if not result.empty:
        result = result.sort_values("Curr Traffic", ascending=False).reset_index(drop=True)
    return result


# ---------------------------------------------------------------------------
# Sheet formatting helpers
# ---------------------------------------------------------------------------

DELTA_COLS = [
    "Score Delta",
    "KW Delta", "Top 3 Delta", "Top 10 Delta", "Top 20 Delta",
    "Traffic Delta", "Volume Delta",
    "Top 3 Vol Delta", "Top 10 Vol Delta", "Top 20 Vol Delta",
]

FILL_HEADER   = PatternFill("solid", fgColor="1F3864")
FILL_SECTION  = PatternFill("solid", fgColor="2E75B6")
FILL_GREEN_H  = PatternFill("solid", fgColor="375623")
FILL_INPUT    = PatternFill("solid", fgColor="FFE699")  # yellow — user input cell
FILL_CF_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED_H    = PatternFill("solid", fgColor="7B2C2C")
FILL_COL_H    = PatternFill("solid", fgColor="D9E1F2")
FILL_WIN_H    = PatternFill("solid", fgColor="E2EFDA")
FILL_LOSE_H   = PatternFill("solid", fgColor="FCE4D6")
FILL_CF_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_CF_RED   = PatternFill("solid", fgColor="FFC7CE")

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BOLD       = Font(bold=True)


def _apply_delta_cf(ws, col_letter: str, n_rows: int) -> None:
    ref = f"{col_letter}2:{col_letter}{n_rows + 1}"
    ws.conditional_formatting.add(
        ref, CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_CF_GREEN)
    )
    ws.conditional_formatting.add(
        ref, CellIsRule(operator="lessThan", formula=["0"], fill=FILL_CF_RED)
    )


def format_url_report_sheet(ws, n_rows: int, delta_col_indices: list[int]) -> None:
    # Bold + colour header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Freeze header, add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    ws.column_dimensions["A"].width = 60  # URL
    ws.column_dimensions["B"].width = 13  # Is Homepage
    for col_idx in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Conditional formatting on delta columns
    for col_idx in delta_col_indices:
        _apply_delta_cf(ws, get_column_letter(col_idx), n_rows)


# ---------------------------------------------------------------------------
# Highlights sheet writer
# ---------------------------------------------------------------------------

def _build_kw_traffic_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build a keyword-level dataframe with traffic change for highlights."""
    col_map = {str(c).strip().lower(): c for c in df.columns}

    kw_col       = find_col(col_map, "Keyword")
    vol_col      = find_col(col_map, "Volume")
    prev_tr_col  = find_col(col_map, "Previous organic traffic")
    curr_tr_col  = find_col(col_map, "Current organic traffic")
    prev_pos_col = find_col(col_map, "Previous position")
    curr_pos_col = find_col(col_map, "Current position")
    prev_url_col = find_col(col_map, "Previous URL")
    curr_url_col = find_col(col_map, "Current URL")

    prev_tr  = pd.to_numeric(df[prev_tr_col], errors="coerce").fillna(0).astype(int)
    curr_tr  = pd.to_numeric(df[curr_tr_col], errors="coerce").fillna(0).astype(int)
    prev_pos = df[prev_pos_col].apply(coerce_position)
    curr_pos = df[curr_pos_col].apply(coerce_position)

    # Prefer current URL; fall back to previous URL for lost rankings
    url = df[curr_url_col].fillna(df[prev_url_col])

    pos_change = [
        (p - c) if (p is not None and c is not None) else None
        for p, c in zip(prev_pos, curr_pos)
    ]

    kw_df = pd.DataFrame({
        "Keyword":        df[kw_col].values,
        "URL":            url.values,
        "Volume":         pd.to_numeric(df[vol_col], errors="coerce").fillna(0).astype(int).values,
        "Prev Position":  df[prev_pos_col].values,
        "Curr Position":  df[curr_pos_col].values,
        "Pos Change":     pos_change,
        "Prev Traffic":   prev_tr.values,
        "Curr Traffic":   curr_tr.values,
        "Traffic Change": (curr_tr - prev_tr).values,
    })
    return kw_df


def write_highlights_sheet(ws, report: pd.DataFrame, df: pd.DataFrame) -> None:
    non_home = report[~report["Is Homepage"]]

    # --- Single highlights ---
    highlights = []
    if not report.empty:
        def pick(df, col, largest=True):
            fn = df[col].idxmax if largest else df[col].idxmin
            try:
                idx = fn()
                return df.loc[idx, "URL"], df.loc[idx, col]
            except Exception:
                return "—", 0

        url, val = pick(report, "Traffic Delta", largest=True)
        highlights.append(("Page with Most Traffic Growth",          url, f"{val:+,}"))

        url, val = pick(report, "Traffic Delta", largest=False)
        highlights.append(("Page with Most Traffic Loss",            url, f"{val:+,}"))

        url, val = pick(report, "Score Delta", largest=True)
        highlights.append(("Most Ranking Score Growth",               url, f"{val:+,}"))

        url, val = pick(report, "Score Delta", largest=False)
        highlights.append(("Most Ranking Score Loss",                 url, f"{val:+,}"))

        url, val = pick(report, "New Rankings", largest=True)
        highlights.append(("Most New Keyword Rankings",              url, f"{val:,}"))

        url, val = pick(report, "Curr Top 3", largest=True)
        highlights.append(("Top Ranking Page Overall (most Top 3 KWs)", url, f"{val:,} Top 3 KWs"))

        if not non_home.empty:
            url, val = pick(non_home, "Curr Traffic", largest=True)
            highlights.append(("Top Traffic Page (excl. Homepage)",  url, f"{val:,}"))

    row = 1

    # Master title
    ws.cell(row=row, column=1, value="KEYWORD RANKINGS CHANGE — PAGE REPORT HIGHLIGHTS")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=row, column=1).fill = FILL_HEADER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # Single Page Highlights section
    ws.cell(row=row, column=1, value="SINGLE PAGE HIGHLIGHTS")
    ws.cell(row=row, column=1).font = FONT_WHITE_BOLD
    ws.cell(row=row, column=1).fill = FILL_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    for col, label in enumerate(["Metric", "URL", "Value"], start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = FONT_BOLD
        c.fill = FILL_COL_H
    row += 1

    for metric, url, value in highlights:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=url)
        ws.cell(row=row, column=3, value=value)
        row += 1

    row += 1  # blank

    # Top 10 Winners
    ws.cell(row=row, column=1, value="TOP 10 WINNERS — RANKING SCORE GROWTH")
    ws.cell(row=row, column=1).font = FONT_WHITE_BOLD
    ws.cell(row=row, column=1).fill = FILL_GREEN_H
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    winners_cols = ["URL", "Prev Score", "Curr Score", "Score Delta", "Prev Top 3", "Curr Top 3", "Traffic Delta", "New Rankings"]
    for col, label in enumerate(winners_cols, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = FONT_BOLD
        c.fill = FILL_WIN_H
    row += 1

    top10 = report.nlargest(10, "Score Delta")
    for _, r in top10.iterrows():
        for col, key in enumerate(winners_cols, start=1):
            ws.cell(row=row, column=col, value=r[key])
        row += 1

    row += 1  # blank

    # Top 10 Losers
    ws.cell(row=row, column=1, value="TOP 10 LOSERS — RANKING SCORE LOSS")
    ws.cell(row=row, column=1).font = FONT_WHITE_BOLD
    ws.cell(row=row, column=1).fill = FILL_RED_H
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    losers_cols = ["URL", "Prev Score", "Curr Score", "Score Delta", "Prev Top 3", "Curr Top 3", "Traffic Delta", "Lost Rankings"]
    for col, label in enumerate(losers_cols, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = FONT_BOLD
        c.fill = FILL_LOSE_H
    row += 1

    bottom10 = report.nsmallest(10, "Score Delta")
    for _, r in bottom10.iterrows():
        for col, key in enumerate(losers_cols, start=1):
            ws.cell(row=row, column=col, value=r[key])
        row += 1

    # --- Keyword-level traffic sections ---
    kw_df = _build_kw_traffic_df(df)
    KW_COLS = ["Keyword", "URL", "Volume", "Prev Position", "Curr Position",
               "Pos Change", "Prev Traffic", "Curr Traffic", "Traffic Change"]
    N_KW_COLS = len(KW_COLS)

    FILL_GAIN_H = PatternFill("solid", fgColor="1E4620")   # dark green
    FILL_LOSS_H = PatternFill("solid", fgColor="7B2C2C")   # dark red

    for section_label, fill_h, fill_row, sort_largest in (
        ("TOP 10 MOST IMPORTANT KEYWORD TRAFFIC GAINS",  FILL_GAIN_H, FILL_WIN_H,  True),
        ("TOP 10 MOST IMPORTANT KEYWORD TRAFFIC LOSSES", FILL_LOSS_H, FILL_LOSE_H, False),
    ):
        row += 1  # blank separator

        ws.cell(row=row, column=1, value=section_label)
        ws.cell(row=row, column=1).font = FONT_WHITE_BOLD
        ws.cell(row=row, column=1).fill = fill_h
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_KW_COLS)
        row += 1

        for col, label in enumerate(KW_COLS, start=1):
            c = ws.cell(row=row, column=col, value=label)
            c.font = FONT_BOLD
            c.fill = fill_row
        row += 1

        top10_kw = (
            kw_df.nlargest(10, "Traffic Change")
            if sort_largest
            else kw_df.nsmallest(10, "Traffic Change")
        )
        for _, r in top10_kw.iterrows():
            for col, key in enumerate(KW_COLS, start=1):
                ws.cell(row=row, column=col, value=r[key])
            row += 1

    # Column widths for highlights sheet
    ws.column_dimensions["A"].width = 35   # Keyword / Metric
    ws.column_dimensions["B"].width = 60   # URL
    ws.column_dimensions["C"].width = 16   # Value / Volume
    for letter in ("D", "E", "F", "G", "H", "I"):
        ws.column_dimensions[letter].width = 15


# ---------------------------------------------------------------------------
# Drilldown helper-tab data builder
# ---------------------------------------------------------------------------

def build_drilldown_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pre-compute per-URL keyword tables for the drilldown tab.
    Returns a flat lookup table keyed by "url|section|rank".
    Sections: top_traffic (top 10), gain (top 5), loss (top 5).
    """
    col_map = {str(c).strip().lower(): c for c in df.columns}

    kw_col       = find_col(col_map, "Keyword")
    vol_col      = find_col(col_map, "Volume")
    prev_tr_col  = find_col(col_map, "Previous organic traffic")
    curr_tr_col  = find_col(col_map, "Current organic traffic")
    prev_pos_col = find_col(col_map, "Previous position")
    curr_pos_col = find_col(col_map, "Current position")
    prev_url_col = find_col(col_map, "Previous URL")
    curr_url_col = find_col(col_map, "Current URL")

    prev_tr_s   = pd.to_numeric(df[prev_tr_col], errors="coerce").fillna(0).astype(int)
    curr_tr_s   = pd.to_numeric(df[curr_tr_col], errors="coerce").fillna(0).astype(int)
    tr_change_s = curr_tr_s - prev_tr_s

    work = pd.DataFrame({
        "keyword":   df[kw_col].values,
        "volume":    pd.to_numeric(df[vol_col], errors="coerce").fillna(0).astype(int).values,
        "prev_pos":  df[prev_pos_col].values,
        "curr_pos":  df[curr_pos_col].values,
        "prev_tr":   prev_tr_s.values,
        "curr_tr":   curr_tr_s.values,
        "tr_change": tr_change_s.values,
        "curr_url":  df[curr_url_col].apply(normalize_url).values,
        "prev_url":  df[prev_url_col].apply(normalize_url).values,
    })

    all_urls = sorted(
        set(work.loc[work["curr_url"] != "", "curr_url"]) |
        set(work.loc[work["prev_url"] != "", "prev_url"])
    )

    OUT_COLS = ["Key", "Keyword", "Volume", "Prev Position", "Curr Position",
                "Prev Traffic", "Curr Traffic", "Traffic Change"]
    rows = []

    for url in all_urls:
        curr_sub = work[work["curr_url"] == url]
        prev_sub = work[work["prev_url"] == url]

        for rank, (_, r) in enumerate(curr_sub.nlargest(10, "curr_tr").iterrows(), start=1):
            rows.append((f"{url}|top_traffic|{rank}",
                         r["keyword"], r["volume"], r["prev_pos"], r["curr_pos"],
                         r["prev_tr"], r["curr_tr"], r["tr_change"]))

        for rank, (_, r) in enumerate(curr_sub.nlargest(5, "tr_change").iterrows(), start=1):
            rows.append((f"{url}|gain|{rank}",
                         r["keyword"], r["volume"], r["prev_pos"], r["curr_pos"],
                         r["prev_tr"], r["curr_tr"], r["tr_change"]))

        for rank, (_, r) in enumerate(prev_sub.nsmallest(5, "tr_change").iterrows(), start=1):
            rows.append((f"{url}|loss|{rank}",
                         r["keyword"], r["volume"], r["prev_pos"], r["curr_pos"],
                         r["prev_tr"], r["curr_tr"], r["tr_change"]))

    return pd.DataFrame(rows, columns=OUT_COLS)


# ---------------------------------------------------------------------------
# URL Drilldown dynamic sheet
# ---------------------------------------------------------------------------

def write_drilldown_sheet(ws, report: pd.DataFrame, n_drilldown_rows: int) -> None:
    """
    Write the dynamic drilldown tab. B3 is the URL input cell.
    Scorecard uses XLOOKUP against URL Report.
    Keyword tables use per-row XLOOKUP against the hidden Drilldown Data tab —
    no spilled array formulas, so openpyxl serialises the file cleanly.

    Bounded row ranges (not full columns) are used in all XLOOKUPs to prevent
    Excel from prepending the implicit-intersection operator '@', which breaks lookups.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    rpt_idx = {c: i + 1 for i, c in enumerate(report.columns)}
    rpt_last = len(report) + 1  # header row + data rows

    def rpt_letter(name):
        return get_column_letter(rpt_idx[name])

    def xlookup(col_name):
        letter = rpt_letter(col_name)
        return (
            f"IF(TRIM($B$3)=\"\",\"—\","
            f"IFERROR(INDEX('URL Report'!${letter}$1:${letter}${rpt_last},"
            f"MATCH(LOWER(TRIM($B$3)),'URL Report'!$A$1:$A${rpt_last},0)),\"—\"))"
        )

    # Helper tab columns: A=Key B=Keyword C=Volume D=PrevPos E=CurrPos
    #                     F=PrevTraffic G=CurrTraffic H=TrafficChange
    HELPER_COLS = ["B", "C", "D", "E", "F", "G", "H"]
    dd_last = n_drilldown_rows + 1  # header row + data rows

    def helper_xlookup(section, rank, col_letter):
        return (
            f"IFERROR(INDEX('Drilldown Data'!${col_letter}$1:${col_letter}${dd_last},"
            f"MATCH(LOWER(TRIM($B$3))&\"|{section}|{rank}\",'Drilldown Data'!$A$1:$A${dd_last},0)),\"\")"
        )

    hp = report[report["Is Homepage"]]
    default_url = hp.iloc[0]["URL"] if not hp.empty else (
        report.iloc[0]["URL"] if not report.empty else ""
    )

    row = 1

    # Title
    c = ws.cell(row=row, column=1, value="URL PERFORMANCE DRILLDOWN")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = FILL_HEADER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 2

    # URL input
    ws.cell(row=row, column=1, value="URL:").font = FONT_BOLD
    inp = ws.cell(row=row, column=2, value=default_url)
    inp.fill = FILL_INPUT
    inp.font = Font(bold=True)
    inp.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    hint = ws.cell(row=row, column=6, value="← select from dropdown or paste URL (lowercase)")
    hint.font = Font(italic=True, color="595959")

    dv = DataValidation(
        type="list",
        formula1=f"'URL Report'!$A$2:$A${len(report) + 2}",
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(ws["B3"])
    row += 2

    # Scorecard
    c = ws.cell(row=row, column=1, value="SCORECARD")
    c.font = FONT_WHITE_BOLD
    c.fill = FILL_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    for col, label in enumerate(["Metric", "Previous", "Current", "Delta"], start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = FONT_BOLD
        c.fill = FILL_COL_H
    row += 1

    scorecard_start = row
    for metric, prev_c, curr_c, delta_c in [
        ("Traffic (est.)",   "Prev Traffic", "Curr Traffic", "Traffic Delta"),
        ("Keywords Ranking", "Prev KWs",     "Curr KWs",     "KW Delta"),
        ("Top 3 Keywords",   "Prev Top 3",   "Curr Top 3",   "Top 3 Delta"),
        ("Ranking Score",    "Prev Score",   "Curr Score",   "Score Delta"),
    ]:
        ws.cell(row=row, column=1, value=metric).font = FONT_BOLD
        ws.cell(row=row, column=2, value=f"={xlookup(prev_c)}")
        ws.cell(row=row, column=3, value=f"={xlookup(curr_c)}")
        ws.cell(row=row, column=4, value=f"={xlookup(delta_c)}")
        row += 1
    scorecard_end = row - 1

    sc_range = f"D{scorecard_start}:D{scorecard_end}"
    ws.conditional_formatting.add(sc_range, CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_CF_GREEN))
    ws.conditional_formatting.add(sc_range, CellIsRule(operator="lessThan",    formula=["0"], fill=FILL_CF_RED))
    ws.conditional_formatting.add(sc_range, CellIsRule(operator="equal",       formula=["0"], fill=FILL_CF_YELLOW))
    row += 1

    KW_HEADERS = [
        "Keyword", "Volume", "Prev Position", "Curr Position",
        "Prev Traffic", "Curr Traffic", "Traffic Change",
    ]

    def write_kw_section(label, hdr_fill, col_fill, section, take_n):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = FONT_WHITE_BOLD
        c.fill = hdr_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        for col, h in enumerate(KW_HEADERS, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = FONT_BOLD
            c.fill = col_fill
        row += 1
        data_start = row
        for rank in range(1, take_n + 1):
            for col_idx, col_letter in enumerate(HELPER_COLS, start=1):
                ws.cell(row=row, column=col_idx, value=f"={helper_xlookup(section, rank, col_letter)}")
            row += 1
        cf_range = f"G{data_start}:G{data_start + take_n - 1}"
        ws.conditional_formatting.add(cf_range, CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_CF_GREEN))
        ws.conditional_formatting.add(cf_range, CellIsRule(operator="lessThan",    formula=["0"], fill=FILL_CF_RED))
        ws.conditional_formatting.add(cf_range, CellIsRule(operator="equal",       formula=["0"], fill=FILL_CF_YELLOW))
        row += 1

    write_kw_section("TOP 10 KEYWORDS BY CURRENT TRAFFIC", FILL_SECTION, FILL_COL_H, "top_traffic", 10)
    write_kw_section("TOP 5 TRAFFIC GAINS IN PERIOD",      FILL_GREEN_H, FILL_WIN_H,  "gain",        5)
    write_kw_section("TOP 5 TRAFFIC LOSSES IN PERIOD",     FILL_RED_H,   FILL_LOSE_H, "loss",        5)

    ws.column_dimensions["A"].width = 42
    for letter in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[letter].width = 15
    ws.column_dimensions["G"].width = 18
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    base_dir = os.environ.get(ENV_DIR_KEY)
    if not base_dir:
        raise EnvironmentError(
            f'Environment variable "{ENV_DIR_KEY}" is not set. '
            f"Please set it to the directory containing {INPUT_BASENAME}.csv"
        )

    input_path = resolve_input_path(base_dir, INPUT_BASENAME)
    if not os.path.isfile(input_path):
        raise FileNotFoundError(
            f"Could not find input file. Expected one of:\n"
            f"  {os.path.join(base_dir, INPUT_BASENAME)}.csv / .xlsx / .xls"
        )

    print(f"Reading: {input_path}")
    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".csv":
        encoding = detect_encoding(input_path)
        print(f"Detected encoding: {encoding}")
        df = pd.read_csv(input_path, encoding=encoding, sep=None, engine="python")
    else:
        df = pd.read_excel(input_path, engine="openpyxl")

    print(f"Columns detected ({len(df.columns)}):")
    for c in df.columns:
        print(f"  {repr(c)}")

    report = build_url_report(df)
    drilldown_data = build_drilldown_data(df)

    output_path = os.path.join(base_dir, OUTPUT_FILENAME)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="URL Report", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Highlights", index=False)
        df.to_excel(writer, sheet_name="Raw Keywords", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="URL Drilldown", index=False)
        drilldown_data.to_excel(writer, sheet_name="Drilldown Data", index=False)

    # Re-open for formatting and highlights content
    wb = load_workbook(output_path)

    # Format URL Report
    ws_report = wb["URL Report"]
    cols = list(report.columns)
    delta_indices = [cols.index(c) + 1 for c in DELTA_COLS if c in cols]
    format_url_report_sheet(ws_report, len(report), delta_indices)

    # Write Highlights
    ws_highlights = wb["Highlights"]
    for row in ws_highlights.iter_rows():
        for cell in row:
            cell.value = None
    write_highlights_sheet(ws_highlights, report, df)

    # Write URL Drilldown
    ws_drilldown = wb["URL Drilldown"]
    for r in ws_drilldown.iter_rows():
        for cell in r:
            cell.value = None
    write_drilldown_sheet(ws_drilldown, report, len(drilldown_data))

    # Hide the helper tab — visible to users who inspect the file but not cluttering the tab bar
    wb["Drilldown Data"].sheet_state = "hidden"

    # Format Raw Keywords — header + freeze + autofilter, URL columns wider
    ws_raw = wb["Raw Keywords"]
    for cell in ws_raw[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = ws_raw.dimensions
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if "url" in str(col_name).lower():
            ws_raw.column_dimensions[letter].width = 55
        elif "keyword" in str(col_name).lower():
            ws_raw.column_dimensions[letter].width = 35
        else:
            ws_raw.column_dimensions[letter].width = 18

    wb.save(output_path)

    print(
        f"Done.\n"
        f"Input:  {input_path}\n"
        f"Output: {output_path}\n"
        f"Sheets: URL Report ({len(report)} pages), Highlights, Raw Keywords ({len(df)} rows), "
        f"URL Drilldown, Drilldown Data ({len(drilldown_data)} rows, hidden)\n"
        f"URLs processed: {len(report)}"
    )


if __name__ == "__main__":
    main()
