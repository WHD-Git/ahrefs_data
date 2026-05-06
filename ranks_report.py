"""
Ahrefs ranking data processor

Input:
- Excel file named "ahrefs_kw_ranks" in the directory set by env var "ahrefs_dir"
  (supports "ahrefs_kw_ranks.xlsx", "ahrefs_kw_ranks.xls", or just "ahrefs_kw_ranks" with .xlsx assumed)

Output:
- Excel file "ahrefs_kw_ranks_processed.xlsx" in the same directory, with 3 sheets:
  1) Raw
  2) Organic
  3) Result Types

Requires:
- pandas
- openpyxl
"""

import os
import re
from urllib.parse import urlparse

from dotenv import load_dotenv
load_dotenv()
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd


# -----------------------------
# Config
# -----------------------------
ENV_DIR_KEY = "ahrefs_dir"
INPUT_BASENAME = "ahrefs_kw_ranks"
OUTPUT_FILENAME = "ahrefs_kw_ranks_processed.xlsx"

# Columns to remove for "Organic" and "Result Types" (exact name matching where possible)
PRUNE_COLUMNS = [
    "CPS",
    "Parent Topic",
    "Parent Topic Volume",
    "Last Update",
    "Referring Domains",
    "Domain rating",
    "Ahrefs Rank",
    "Traffic",
    "Keywords",
    "CPC",
    "Global volume",
    "Traffic potential",
    "Global traffic potential",
    "First seen",
    "Intents",
    "Languages",
    "Page type",
    # These often appear with date suffixes; we’ll prune by prefix match too:
    "SV trend",
    "SV Forecasting trend",
]


# -----------------------------
# Helpers
# -----------------------------
def resolve_input_path(base_dir: str, basename: str) -> str:
    """Resolve input path allowing basename, .xlsx, or .xls."""
    candidates = [
        os.path.join(base_dir, basename),
        os.path.join(base_dir, f"{basename}.xlsx"),
        os.path.join(base_dir, f"{basename}.xls"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    # default to .xlsx if none exist (so error is clear)
    return os.path.join(base_dir, f"{basename}.xlsx")


def normalize_domain(url: str) -> str:
    """Extract registrable-looking domain/host from URL and strip leading www."""
    if url is None:
        return ""
    s = str(url).strip()
    if not s or s.lower() == "nan":
        return ""

    # Ensure scheme so urlparse can read netloc reliably
    if not re.match(r"^[a-zA-Z][a-zA-Z0-9+\-.]*://", s):
        s_for_parse = "http://" + s
    else:
        s_for_parse = s

    parsed = urlparse(s_for_parse)
    host = (parsed.netloc or "").strip()

    # If still empty (some weird formats), fall back to path start
    if not host:
        # e.g. "example.com/page" without scheme sometimes becomes path in urlparse
        host = parsed.path.split("/")[0].strip()

    host = host.lower()
    if host.startswith("www."):
        host = host[4:]

    # Remove credentials/ports if present
    host = host.split("@")[-1].split(":")[0]

    return host


def coerce_position(val):
    """Try to coerce position into int; return None if not possible."""
    if val is None:
        return None
    try:
        # Handle floats like 3.0
        f = float(str(val).strip())
        if pd.isna(f):
            return None
        return int(f)
    except Exception:
        return None


def prune_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Drop explicit columns and any columns that start with date-series prefixes."""
    cols = list(df.columns)

    to_drop = set()

    # Exact drops where they exist
    for c in cols:
        if c in PRUNE_COLUMNS:
            to_drop.add(c)

    # Prefix-based drops for date series columns (e.g. "SV trend (2023-01-01)" etc.)
    prefix_prune = ("SV trend", "SV Forecasting trend")
    for c in cols:
        for pref in prefix_prune:
            if str(c).startswith(pref):
                to_drop.add(c)

    return df.drop(columns=[c for c in cols if c in to_drop], errors="ignore")


def contains_any_type(cell_value: str, needles) -> bool:
    """Case-insensitive 'contains' check for any needle in a possibly multi-valued cell."""
    if cell_value is None:
        return False
    s = str(cell_value).lower()
    return any(n.lower() in s for n in needles)


# -----------------------------
# Domain aggregation
# -----------------------------
FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")


def build_domain_summary(df: pd.DataFrame, pos_col: str, type_col: str = None, type_filter=None) -> pd.DataFrame:
    """Aggregate by Domain. Optionally pre-filter rows by type_filter values in type_col."""
    work = df.copy()
    if type_filter and type_col:
        work = work[work[type_col].apply(lambda v: contains_any_type(v, type_filter))]
    if work.empty:
        return pd.DataFrame(columns=["Domain", "Keywords", "Top 3", "Top 10", "Top 20", "Ranking Score"])

    col_map = {str(c).strip().lower(): c for c in work.columns}
    traffic_col = col_map.get("traffic")
    volume_col = col_map.get("volume")

    pos_int = work[pos_col].apply(coerce_position)
    work = work.copy()
    work["_t3"]  = pos_int.apply(lambda x: 1 if (x is not None and x <= 3)  else 0)
    work["_t10"] = pos_int.apply(lambda x: 1 if (x is not None and x <= 10) else 0)
    work["_t20"] = pos_int.apply(lambda x: 1 if (x is not None and x <= 20) else 0)

    agg = work.groupby("Domain", as_index=False).agg(
        Keywords=("Domain", "count"),
        Top_3=("_t3",  "sum"),
        Top_10=("_t10", "sum"),
        Top_20=("_t20", "sum"),
    )

    if traffic_col:
        tr = work.groupby("Domain")[traffic_col].sum().rename("Traffic")
        agg = agg.merge(tr, on="Domain", how="left")
    if volume_col:
        vol = work.groupby("Domain")[volume_col].sum().rename("Volume")
        agg = agg.merge(vol, on="Domain", how="left")

    agg["Ranking Score"] = (
        agg["Top_3"] * 3
        + (agg["Top_10"] - agg["Top_3"]) * 2
        + (agg["Top_20"] - agg["Top_10"]) * 1
    )
    agg = agg.rename(columns={"Top_3": "Top 3", "Top_10": "Top 10", "Top_20": "Top 20"})

    ordered = ["Domain", "Keywords"]
    if traffic_col:
        ordered.append("Traffic")
    if volume_col:
        ordered.append("Volume")
    ordered += ["Top 3", "Top 10", "Top 20", "Ranking Score"]
    agg = agg[[c for c in ordered if c in agg.columns]]

    return agg.sort_values("Ranking Score", ascending=False).reset_index(drop=True)


def build_domain_by_type(df: pd.DataFrame, pos_col: str, type_col: str) -> pd.DataFrame:
    """Aggregate by Domain + Type so the sheet can be filtered by result type."""
    work = df.copy()
    col_map = {str(c).strip().lower(): c for c in work.columns}
    traffic_col = col_map.get("traffic")
    volume_col = col_map.get("volume")

    pos_int = work[pos_col].apply(coerce_position)
    work = work.copy()
    work["_t3"]  = pos_int.apply(lambda x: 1 if (x is not None and x <= 3)  else 0)
    work["_t10"] = pos_int.apply(lambda x: 1 if (x is not None and x <= 10) else 0)
    work["_t20"] = pos_int.apply(lambda x: 1 if (x is not None and x <= 20) else 0)

    grp = ["Domain", type_col]
    agg = work.groupby(grp, as_index=False).agg(
        Keywords=(type_col, "count"),
        Top_3=("_t3",  "sum"),
        Top_10=("_t10", "sum"),
        Top_20=("_t20", "sum"),
    )

    if traffic_col:
        tr = work.groupby(grp)[traffic_col].sum().reset_index().rename(columns={traffic_col: "Traffic"})
        agg = agg.merge(tr, on=grp, how="left")
    if volume_col:
        vol = work.groupby(grp)[volume_col].sum().reset_index().rename(columns={volume_col: "Volume"})
        agg = agg.merge(vol, on=grp, how="left")

    agg["Ranking Score"] = (
        agg["Top_3"] * 3
        + (agg["Top_10"] - agg["Top_3"]) * 2
        + (agg["Top_20"] - agg["Top_10"]) * 1
    )
    agg = agg.rename(columns={"Top_3": "Top 3", "Top_10": "Top 10", "Top_20": "Top 20", type_col: "Type"})

    ordered = ["Domain", "Type", "Keywords"]
    if traffic_col:
        ordered.append("Traffic")
    if volume_col:
        ordered.append("Volume")
    ordered += ["Top 3", "Top 10", "Top 20", "Ranking Score"]
    agg = agg[[c for c in ordered if c in agg.columns]]

    return agg.sort_values(["Ranking Score"], ascending=False).reset_index(drop=True)


def format_summary_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


# -----------------------------
# Main processing
# -----------------------------
def main():
    base_dir = os.environ.get(ENV_DIR_KEY)
    if not base_dir:
        raise EnvironmentError(
            f'Environment variable "{ENV_DIR_KEY}" is not set. '
            f"Please set it to the directory containing {INPUT_BASENAME}.xlsx"
        )

    input_path = resolve_input_path(base_dir, INPUT_BASENAME)
    if not os.path.isfile(input_path):
        raise FileNotFoundError(
            f"Could not find input file. Looked for:\n"
            f"- {os.path.join(base_dir, INPUT_BASENAME)}\n"
            f"- {os.path.join(base_dir, INPUT_BASENAME + '.xlsx')}\n"
            f"- {os.path.join(base_dir, INPUT_BASENAME + '.xls')}\n"
            f"\nResolved default path would be: {input_path}"
        )

    # Read first sheet by default
    xls = pd.ExcelFile(input_path)
    sheet_name = xls.sheet_names[0]
    df = pd.read_excel(input_path, sheet_name=sheet_name, engine="openpyxl")

    # Standardize column lookups (preserve original names but find key columns robustly)
    col_map_lower = {str(c).strip().lower(): c for c in df.columns}

    url_col = col_map_lower.get("url")
    pos_col = col_map_lower.get("position")
    type_col = col_map_lower.get("type")

    if not url_col:
        raise KeyError('Required column "URL" not found (case-insensitive).')
    if not pos_col:
        raise KeyError('Required column "Position" not found (case-insensitive).')
    if not type_col:
        raise KeyError('Required column "Type" not found (case-insensitive).')

    # -------------------------
    # 1) RAW
    # -------------------------
    raw = df.copy()

    # Domain column next to URL
    raw["Domain"] = raw[url_col].apply(normalize_domain)
    # Insert Domain right after URL
    cols = list(raw.columns)
    cols.remove("Domain")
    url_idx = cols.index(url_col)
    cols.insert(url_idx + 1, "Domain")
    raw = raw[cols]

    # Top 3 / Top 10 columns based on numeric position
    pos_int = raw[pos_col].apply(coerce_position)
    raw["Top 3"] = pos_int.apply(lambda x: 1 if (x is not None and x < 4) else 0)
    raw["Top 10"] = pos_int.apply(lambda x: 1 if (x is not None and x < 11) else 0)

    # -------------------------
    # 2) ORGANIC
    # -------------------------
    organic = raw.copy()

    # Keep only rows where URL is not empty
    organic = organic[organic[url_col].notna() & (organic[url_col].astype(str).str.strip() != "")]

    # Filter Type contains "Organic" OR "AI Overview"
    organic = organic[organic[type_col].apply(lambda v: contains_any_type(v, ["Organic", "AI Overview"]))]

    # Prune columns
    organic = prune_columns(organic)

    # -------------------------
    # 3) RESULT TYPES (no Type filter)
    # -------------------------
    result_types = raw.copy()
    result_types = result_types[
        result_types[url_col].notna() & (result_types[url_col].astype(str).str.strip() != "")
    ]
    result_types = prune_columns(result_types)

    # -------------------------
    # 4) DOMAIN SUMMARY (all result types)
    # -------------------------
    domain_summary = build_domain_summary(raw, pos_col)

    # -------------------------
    # 5) DOMAIN BY TYPE (pivot: domain × result type)
    # -------------------------
    domain_by_type = build_domain_by_type(raw, pos_col, type_col)

    # -------------------------
    # 6) BLUE LINKS (organic / AI overview only, by domain)
    # -------------------------
    blue_links = build_domain_summary(
        raw, pos_col, type_col=type_col, type_filter=["Organic", "AI Overview"]
    )

    # -------------------------
    # Write output
    # -------------------------
    output_path = os.path.join(base_dir, OUTPUT_FILENAME)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        raw.to_excel(writer, sheet_name="Raw", index=False)
        organic.to_excel(writer, sheet_name="Organic", index=False)
        result_types.to_excel(writer, sheet_name="Result Types", index=False)
        domain_summary.to_excel(writer, sheet_name="Domain Summary", index=False)
        domain_by_type.to_excel(writer, sheet_name="Domain by Type", index=False)
        blue_links.to_excel(writer, sheet_name="Blue Links", index=False)

    # Apply formatting to analysis tabs
    wb = load_workbook(output_path)
    for sheet_name in ("Domain Summary", "Domain by Type", "Blue Links"):
        format_summary_sheet(wb[sheet_name])
    wb.save(output_path)

    print(
        f"Done.\nInput:  {input_path}\nOutput: {output_path}\n"
        f"Sheets: Raw, Organic, Result Types, Domain Summary, Domain by Type, Blue Links"
    )


if __name__ == "__main__":
    main()